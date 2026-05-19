from db.context import get_sync_engine, get_async_engine
from sqlalchemy import Delete
from sqlalchemy import Update
from sqlalchemy import or_
from sqlalchemy import and_
from sqlalchemy import case
from sqlalchemy import func
from sqlalchemy.orm import sessionmaker, aliased
from sqlalchemy.sql import func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from models.db import IvrCampaigns, CampaignNumbers, IvrCarriers, IvrFlows, IvrCallerIds
from models.dto import IvrCampaignsModel, IvrCarriersModel, IvrFlowsModel
from fastapi.responses import JSONResponse
from fastapi import status
from openpyxl import load_workbook
import pandas as pd
from typing import Optional, List

def create(campaignname: str, campaigndescription: str, carrierid: int, carriername: str, flowid: int, flowname: str, database: str):
    sync_engine = get_sync_engine(database)
    sync_session_maker = sessionmaker(bind=sync_engine, expire_on_commit=False)
    session = sync_session_maker()
    try:
        ivrcampaign = IvrCampaigns()
        ivrcampaign.i_campaignName = campaignname
        ivrcampaign.i_campaignDescription = campaigndescription
        ivrcampaign.i_carrierId = carrierid
        ivrcampaign.i_carrierName = carriername
        ivrcampaign.i_flowId = flowid
        ivrcampaign.i_flowName = flowname
        session.add(ivrcampaign)
        try:
            session.commit()
            session.close()
            return ivrcampaign.i_campaignId
        except IntegrityError as e:
            session.rollback()
            session.close()
            return str(e.orig)
    finally:
        sync_engine.dispose()

async def numbersCreate(file_path: str, file_extension: str, campaignid: int, campaignname: str, campaigndescription: str, database: str):
    async_engine = get_async_engine(database)
    async_session_maker = sessionmaker(async_engine, expire_on_commit=False, class_=AsyncSession)
    
    async with async_session_maker() as session:
        try:
            chunk_size = 10_000  # Process records in batches
            
            if file_extension == 'csv':
                reader = pd.read_csv(file_path, chunksize=chunk_size)
                for chunk in reader:
                    records = [
                        {
                            "c_campaignId": campaignid,
                            "c_campaignName": campaignname,
                            "c_campaignDescription": campaigndescription,
                            "c_campaignNumber": row.iloc[0],
                        }
                        for _, row in chunk.iterrows()
                    ]
                    if records:
                        await session.execute(CampaignNumbers.__table__.insert(), records)
                        await session.commit()

            elif file_extension in ['xls', 'xlsx']:
                wb = load_workbook(file_path, read_only=True)  # Read-only mode for performance
                ws = wb.active  # Get the first sheet
                
                records = []
                for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if index == 1:  # Skip the header row
                        continue
                    records.append({
                        "c_campaignId": campaignid,
                        "c_campaignName": campaignname,
                        "c_campaignDescription": campaigndescription,
                        "c_campaignNumber": row[0],
                    })
                    if len(records) >= chunk_size:
                        await session.execute(CampaignNumbers.__table__.insert(), records)
                        await session.commit()
                        records.clear()  # Reset batch buffer

                # Insert remaining records
                if records:
                    await session.execute(CampaignNumbers.__table__.insert(), records)
                    await session.commit()

            return "success"

        except Exception as e:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"message": f"Error Processing File: {str(e)}"}
            )
        finally:
            await async_engine.dispose()

async def update(campaignid: int, campaignname: str, campaigndescription: str, carrierid: int, carriername: str, flowid: int, flowname: str, database: str):
    async_engine = get_async_engine(database)
    async_session_maker = sessionmaker(async_engine, expire_on_commit=False, class_=AsyncSession)
    session = async_session_maker()
    try:
        await session.execute(Update(IvrCampaigns).where(IvrCampaigns.i_campaignId == campaignid).values({
            IvrCampaigns.i_campaignName: campaignname,
            IvrCampaigns.i_campaignDescription: campaigndescription,
            IvrCampaigns.i_carrierId: carrierid,
            IvrCampaigns.i_carrierName: carriername,
            IvrCampaigns.i_flowId: flowid,
            IvrCampaigns.i_flowName: flowname
        }))
        await session.execute(Update(CampaignNumbers).where(CampaignNumbers.c_campaignId == campaignid).values({
            CampaignNumbers.c_campaignName: campaignname,
            CampaignNumbers.c_campaignDescription: campaigndescription
        }))
        await session.commit()
        await session.close()
    finally:
        await async_engine.dispose()

async def delete(campaignid: int, database: str):
    async_engine = get_async_engine(database)
    async_session_maker = sessionmaker(async_engine, expire_on_commit=False, class_=AsyncSession)
    session = async_session_maker()
    try:
        await session.execute(Delete(IvrCampaigns).where(
                and_(
                    IvrCampaigns.i_campaignId == campaignid
                )
            )
        )
        await session.execute(Delete(CampaignNumbers).where(
                and_(
                    CampaignNumbers.c_campaignId == campaignid
                )
            )
        )
        await session.commit()
        await session.close()
    finally:
        await async_engine.dispose()

def fetch(database: str, limit: int = 1000, offset: int = 0, searchString: str = "") -> dict:
    sync_engine = get_sync_engine(database)
    sync_session_maker = sessionmaker(bind=sync_engine, expire_on_commit=False)
    session = sync_session_maker()
    try:
        # Base query for campaigns
        recordQuery = session.query(
            IvrCampaigns,
            func.count(CampaignNumbers.c_campaignId).label("c_totalLeads"),
            func.count(case((CampaignNumbers.c_status != 'pending', 1), else_=None)).label("c_completedLeads")
        ).outerjoin(CampaignNumbers, IvrCampaigns.i_campaignId == CampaignNumbers.c_campaignId).group_by(IvrCampaigns.i_campaignId)

        # Apply search filters if provided
        if searchString:
            recordQuery = recordQuery.filter(
                or_(
                    IvrCampaigns.i_campaignId.like(f"%{searchString}%"),
                    IvrCampaigns.i_campaignName.like(f"%{searchString}%"),
                    IvrCampaigns.i_campaignDescription.like(f"%{searchString}%")
                )
            )

        # Get total count after filtering
        totalRecordsCount = recordQuery.count()

        # Paginate the results
        totalRecordsUnserialized = recordQuery.limit(limit).offset(offset).all()

        # Format the response
        totalRecords = [
            {
                **IvrCampaignsModel.from_orm(record[0]).dict(),
                "c_totalLeads": record[1],
                "c_completedLeads": record[2]
            }
            for record in totalRecordsUnserialized
        ]

        return {
            "totalRecordsCount": totalRecordsCount,
            "totalRecords": totalRecords
        }
    finally:
        session.close()
        sync_engine.dispose()

def listCarrier(database: str) -> dict:
    sync_engine = get_sync_engine(database)
    sync_session_maker = sessionmaker(bind=sync_engine, expire_on_commit=False)
    session = sync_session_maker()
    try:
        recordQuery = session.query(IvrCarriers).filter()
        totalRecordsUnserialized = recordQuery.all()
        totalRecords = [IvrCarriersModel.from_orm(record).dict() for record in totalRecordsUnserialized]
        return {
            "totalRecords": totalRecords
        }
    finally:
        session.close()
        sync_engine.dispose()

def listFlow(database: str) -> dict:
    sync_engine = get_sync_engine(database)
    sync_session_maker = sessionmaker(bind=sync_engine, expire_on_commit=False)
    session = sync_session_maker()
    try:
        recordQuery = session.query(IvrFlows).filter()
        totalRecordsUnserialized = recordQuery.all()
        totalRecords = [IvrFlowsModel.from_orm(record).dict() for record in totalRecordsUnserialized]
        return {
            "totalRecords": totalRecords
        }
    finally:
        session.close()
        sync_engine.dispose()

def check(database: str, campaignname: str):
    sync_engine = get_sync_engine(database)
    sync_session_maker = sessionmaker(bind=sync_engine, expire_on_commit=False)
    session = sync_session_maker()
    try:
        existingCampaignName = session.query(IvrCampaigns).filter(IvrCampaigns.i_campaignName == campaignname).first()
        errors = []
        if existingCampaignName:
            errors.append(f"CampaignName '{campaignname}' already exists.")
        if errors:
            session.close()
            return {"uniqueConstraint":"Yes", "data": errors}
        session.close()
        return {"uniqueConstraint":"No"}
    finally:
        sync_engine.dispose()

def listCallerId(database: str) -> dict:
    sync_engine = get_sync_engine(database)
    sync_session_maker = sessionmaker(bind=sync_engine, expire_on_commit=False)
    session = sync_session_maker()
    try:
        result = (
            session.query(
                IvrCampaigns.i_campaignId,
                IvrCampaigns.i_campaignName,
                IvrCampaigns.i_carrierName,
                func.group_concat(IvrCallerIds.carrier_name).label("carrier_name")
            )
            .outerjoin(IvrCallerIds, IvrCampaigns.i_campaignId == IvrCallerIds.campaign_id)
            .group_by(
                IvrCampaigns.i_campaignId,
                IvrCampaigns.i_campaignName,
                IvrCampaigns.i_carrierName
            )
            .all()
        )
        totalRecords = [
            {
                "i_campaignId": row.i_campaignId,
                "i_campaignName": row.i_campaignName,
                "i_carrierName": row.i_carrierName,
                "carrier_name": row.carrier_name or ""
            }
            for row in result
        ]
        return {
            "totalRecords": totalRecords
        }
    finally:
        session.close()
        sync_engine.dispose()

async def createCallerId(campaignid: int, callerids: List[int], database: str):
    async_engine = get_async_engine(database)
    async_session_maker = sessionmaker(async_engine, expire_on_commit=False, class_=AsyncSession)
    session = async_session_maker()
    try:
        for callerid in callerids:
            ivrcallerid = IvrCallerIds()
            ivrcallerid.campaign_id = campaignid
            ivrcallerid.carrier_name = callerid
            ivrcallerid.sequence = 0
            ivrcallerid.status = "active"

            session.add(ivrcallerid)
        try:
            await session.commit()
            await session.close()
            return "success"
        except IntegrityError as e:
            await session.rollback()
            await session.close()
            return str(e.orig)
    finally:
        await async_engine.dispose()